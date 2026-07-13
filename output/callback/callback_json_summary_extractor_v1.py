#!/usr/bin/env python3
"""
callback_json_summary_extractor_v1.py

Purpose
-------
Extract all non-leg callback JSON reporting into a reportable Excel workbook.
This complements the separate detailed leg extractor. It deliberately ignores
*_by_leg dictionaries and focuses on scalar summaries, solver bounds, DP totals,
route-DP outcomes, subtour statistics, feasibility-cut statistics, pruning
statistics, optimality-cut statistics, and compact histogram/distribution fields.

Selected run convention
-----------------------
Uses the v2 selected-run set:
- c101C6_1_K1_callback_report.json
- c101C6_2_K1_callback_report.json
- c101C12_1_K1_callback_report.json
- c101C12_2_K1_callback_report.json
- c103C16_1_K1_callback_report.json
- c103C16_2_K3_callback_report.json
- c104C10_1_K1_callback_report.json
- c104C10_2_K1_callback_report.json

C101C6_3 and C101C6_4 are intentionally excluded.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SELECTED_FILES = [
    "c101C6_1_K1_callback_report.json",
    "c101C6_2_K1_callback_report.json",
    "c101C12_1_K1_callback_report.json",
    "c101C12_2_K1_callback_report.json",
    "c103C16_1_K1_callback_report.json",
    "c103C16_2_K3_callback_report.json",
    "c104C10_1_K1_callback_report.json",
    "c104C10_2_K1_callback_report.json",
]

# Exclude detailed leg dictionaries because those are handled by the leg extractor.
EXCLUDE_DICT_SUFFIXES = ("_by_leg",)


def infer_instance_and_k(file_name: str) -> Tuple[str, str]:
    stem = Path(file_name).stem.replace("_callback_report", "")
    m = re.match(r"(?P<instance>.+)_K(?P<k>\d+)$", stem)
    if not m:
        return stem, ""
    return m.group("instance"), f"K{m.group('k')}"


def is_scalar(x: Any) -> bool:
    return x is None or isinstance(x, (str, int, float, bool))


def norm_value(x: Any) -> Any:
    if isinstance(x, bool) or x is None:
        return x
    return x


def flatten_scalars(prefix: str, obj: Dict[str, Any]) -> Dict[str, Any]:
    """Return only scalar fields from a dictionary. Nested dictionaries are ignored here."""
    out: Dict[str, Any] = {}
    for k, v in (obj or {}).items():
        col = f"{prefix}.{k}" if prefix else k
        if is_scalar(v):
            out[col] = norm_value(v)
    return out


def dict_items_for_section(
    instance: str,
    selected_k: str,
    file_name: str,
    section: str,
    obj: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Convert compact distribution dictionaries to long rows.

    Detailed *_by_leg dictionaries are excluded here to avoid duplicating the
    separate detailed leg workbook.
    """
    rows: List[Dict[str, Any]] = []
    for field, val in (obj or {}).items():
        if any(field.endswith(suffix) for suffix in EXCLUDE_DICT_SUFFIXES):
            continue
        if isinstance(val, dict):
            for bucket, count in val.items():
                rows.append({
                    "instance": instance,
                    "selected_k": selected_k,
                    "file_name": file_name,
                    "section": section,
                    "field": field,
                    "bucket": str(bucket),
                    "value": count,
                })
    return rows


def extract_file(file_name: str, data: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]]:
    instance, selected_k = infer_instance_and_k(file_name)
    base = {"instance": instance, "selected_k": selected_k, "file_name": file_name}

    bounds = data.get("bounds", {}) or {}
    solver_bounds = data.get("solver_bounds", {}) or {}
    dp = data.get("dp_internal_summary", {}) or {}
    route_dp = data.get("route_dp_summary", {}) or {}
    opt = data.get("optimality_cut_summary", {}) or {}
    feas = data.get("feasibility_cut_summary", {}) or {}
    prune = data.get("feasibility_cut_pruning_summary", {}) or {}
    subtour = data.get("subtour_cut_summary", {}) or {}
    comp = data.get("subtour_component_summary", {}) or {}

    # Report-facing wide summary: only scalar columns.
    summary = dict(base)
    summary.update({
        "schema_version": data.get("schema_version"),
        "objective_sense": data.get("objective_sense"),
    })
    summary.update(flatten_scalars("bounds", bounds))
    summary.update(flatten_scalars("solver_bounds", solver_bounds))
    summary.update(flatten_scalars("dp", dp))
    summary.update(flatten_scalars("route_dp", route_dp))
    summary.update(flatten_scalars("optimality_cut", opt))
    summary.update(flatten_scalars("feasibility_cut", feas))
    summary.update(flatten_scalars("feasibility_pruning", prune))
    summary.update(flatten_scalars("subtour_cut", subtour))
    summary.update(flatten_scalars("subtour_component", comp))

    # Section-specific scalar tables.
    section_rows: Dict[str, List[Dict[str, Any]]] = {
        "bounds": [{**base, **flatten_scalars("", bounds)}],
        "solver_bounds": [{**base, **flatten_scalars("", solver_bounds)}],
        "dp_internal_summary": [{**base, **flatten_scalars("", dp)}],
        "route_dp_summary": [{**base, **flatten_scalars("", route_dp)}],
        "optimality_cut_summary": [{**base, **flatten_scalars("", opt)}],
        "feasibility_cut_summary": [{**base, **flatten_scalars("", feas)}],
        "feasibility_pruning_summary": [{**base, **flatten_scalars("", prune)}],
        "subtour_cut_summary": [{**base, **flatten_scalars("", subtour)}],
        "subtour_component_summary": [{**base, **flatten_scalars("", comp)}],
    }

    histogram_rows: List[Dict[str, Any]] = []
    for section, obj in [
        ("dp_internal_summary", dp),
        ("route_dp_summary", route_dp),
        ("optimality_cut_summary", opt),
        ("feasibility_cut_summary", feas),
        ("feasibility_cut_pruning_summary", prune),
        ("subtour_cut_summary", subtour),
        ("subtour_component_summary", comp),
    ]:
        histogram_rows.extend(dict_items_for_section(instance, selected_k, file_name, section, obj))

    return summary, section_rows, histogram_rows


def build_validation(summary_df: pd.DataFrame, histogram_df: pd.DataFrame, missing: List[str]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    rows.append({"check": "selected_files_read", "value": int(len(summary_df)), "detail": "Number of selected JSON files successfully read."})
    rows.append({"check": "histogram_rows", "value": int(len(histogram_df)), "detail": "Number of compact dictionary/distribution rows extracted, excluding *_by_leg."})
    for m in missing:
        rows.append({"check": "missing_selected_file", "value": m, "detail": "Expected selected JSON file was not found in input folder."})
    if not summary_df.empty:
        # A simple integrity check for scalar dimensions only.
        for col in ["bounds.upper_bound", "bounds.lower_bound", "bounds.runtime_seconds", "dp.arrival_label_dominated", "dp.station_label_dominated", "route_dp.total_calls", "subtour_cut.subtour_cuts"]:
            rows.append({"check": f"non_null_{col}", "value": int(summary_df[col].notna().sum()) if col in summary_df.columns else 0, "detail": "Count of selected files with this scalar populated."})
    return pd.DataFrame(rows)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 80))
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(max_len + 2, 45))
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=".", help="Folder containing callback JSON files")
    parser.add_argument("--output", default="callback_json_summary_extraction_v1.xlsx", help="Output Excel workbook")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    summaries: List[Dict[str, Any]] = []
    section_acc: Dict[str, List[Dict[str, Any]]] = {}
    histograms: List[Dict[str, Any]] = []
    missing: List[str] = []

    for file_name in SELECTED_FILES:
        path = input_dir / file_name
        if not path.exists():
            missing.append(file_name)
            continue
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        summary, section_rows, histogram_rows = extract_file(file_name, data)
        summaries.append(summary)
        for section, rows in section_rows.items():
            section_acc.setdefault(section, []).extend(rows)
        histograms.extend(histogram_rows)

    summary_df = pd.DataFrame(summaries)
    histogram_df = pd.DataFrame(histograms).sort_values(["instance", "section", "field", "bucket"]) if histograms else pd.DataFrame(columns=["instance", "selected_k", "file_name", "section", "field", "bucket", "value"])

    # Report-friendly column ordering for the wide workbook.
    preferred_first = [
        "instance", "selected_k", "file_name", "schema_version", "objective_sense",
        "bounds.upper_bound", "bounds.lower_bound", "bounds.relative_gap", "bounds.runtime_seconds", "bounds.node_count", "bounds.snapshots",
        "dp.arrival_label_dominated", "dp.arrival_labels_removed_by_new", "dp.route_label_dominated", "dp.route_labels_removed_by_new",
        "dp.station_label_dominated", "dp.station_revisit_blocked", "dp.energy_infeasible_transitions",
        "dp.destination_time_window_rejections", "dp.station_horizon_rejections", "dp.leg_frontier_calls", "dp.leg_frontier_success", "dp.leg_frontier_no_arrivals",
        "route_dp.total_calls", "route_dp.success", "route_dp.failure", "route_dp.skipped_short_skeleton",
        "optimality_cut.total", "optimality_cut.delta_total_count", "optimality_cut.delta_total_sum", "optimality_cut.delta_total_min", "optimality_cut.delta_total_max",
        "feasibility_cut.total", "feasibility_cut.active", "feasibility_cut.fallback",
        "feasibility_pruning.total", "feasibility_pruning.front_prune_success", "feasibility_pruning.front_prune_failure",
        "subtour_cut.mipsol_callbacks", "subtour_cut.subtour_cuts", "subtour_cut.returned_before_route_dp", "subtour_cut.reached_route_dp", "subtour_cut.routes_checked",
        "subtour_component.analysed", "subtour_component.component_cut_mode_used", "subtour_component.aggregate_cut_mode_used",
    ]
    if not summary_df.empty:
        ordered = [c for c in preferred_first if c in summary_df.columns] + [c for c in summary_df.columns if c not in preferred_first]
        summary_df = summary_df[ordered]

    notes = pd.DataFrame([
        {"item": "scope", "detail": "This workbook extracts non-leg JSON reporting only. Detailed *_by_leg dictionaries are intentionally excluded because they belong in the separate leg workbook."},
        {"item": "selected_runs", "detail": ", ".join(SELECTED_FILES)},
        {"item": "excluded_instances", "detail": "C101C6_3 and C101C6_4 are excluded by instruction."},
        {"item": "histogram_long", "detail": "Compact dictionaries such as route_size, fail_i, reason, cut_size, chosen_arc_count and component sizes are normalised to long format."},
        {"item": "feasibility_patterns", "detail": "The feasibility_cut_summary.pattern dictionary is included in histogram_long. It may be large but is preserved for auditability."},
    ])

    sheets: Dict[str, pd.DataFrame] = {
        "run_summary_wide": summary_df,
        "histogram_long": histogram_df,
        "validation": build_validation(summary_df, histogram_df, missing),
        "notes": notes,
    }
    for section, rows in section_acc.items():
        sheets[section] = pd.DataFrame(rows)

    write_excel(Path(args.output), sheets)
    print(f"Wrote {args.output}")
    if missing:
        print("Missing selected JSON files:", ", ".join(missing))


if __name__ == "__main__":
    main()
