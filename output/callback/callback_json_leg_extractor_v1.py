#!/usr/bin/env python3
"""
callback_json_leg_extractor_v1.py

Purpose
-------
Read the selected callback JSON reports, append full by-leg diagnostics to the
existing callback_json_extraction workbook, and write a versioned Excel output.

Expected input
--------------
Place this script in, or run it from, the folder containing the JSON callback
reports. Alternatively pass --input-dir and --base-workbook.

Selected run convention
-----------------------
Uses the v2 selected-run set:
- c101C6_1_K1_callback_report.json
- c101C6_2_K1_callback_report.json
- c101C12_1_K1_callback_report.json
- c101C12_2_K1_callback_report.json
- c103C16_1_K1_callback_report.json
- c103C16_2_K3_callback_report.json
- c104C10_1_K1_callback_report.json
- c104C10_2_K1_callback_report.json

C101C6_3 and C101C6_4 are intentionally excluded.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SELECTED_FILES = [
    "c101C6_1_K1_callback_report.json",
    "c101C6_2_K1_callback_report.json",
    "c101C12_1_K1_callback_report.json",
    "c101C12_2_K1_callback_report.json",
    "c103C16_1_K1_callback_report.json",
    "c103C16_2_K3_callback_report.json",
    "c104C10_1_K1_callback_report.json",
    "c104C10_2_K1_callback_report.json",
]

BY_LEG_METRICS = [
    "arrival_label_dominated_by_leg",
    "energy_infeasible_by_leg",
    "no_arrivals_by_leg",
    "destination_tw_by_leg",
    "station_horizon_by_leg",
    "station_label_dominated_by_leg",
]

TOTAL_FIELDS = {
    "arrival_label_dominated_by_leg": "arrival_label_dominated",
    "energy_infeasible_by_leg": "energy_infeasible_transitions",
    "no_arrivals_by_leg": "leg_frontier_no_arrivals",
    "destination_tw_by_leg": "destination_time_window_rejections",
    "station_horizon_by_leg": "station_horizon_rejections",
    "station_label_dominated_by_leg": "station_label_dominated",
}


def infer_instance_and_k(file_name: str) -> Tuple[str, str]:
    stem = Path(file_name).stem.replace("_callback_report", "")
    m = re.match(r"(?P<instance>.+)_K(?P<k>\d+)$", stem)
    if not m:
        return stem, ""
    return m.group("instance"), f"K{m.group('k')}"


def split_leg(leg: str) -> Tuple[str, str]:
    # Handles both C1->C2 and escaped/spaced variants if they appear.
    if "->" in leg:
        a, b = leg.split("->", 1)
    elif "- >" in leg:
        a, b = leg.split("- >", 1)
    else:
        return leg, ""
    return a.strip(), b.strip()


def flatten_by_leg(file_name: str, data: Dict[str, Any]) -> List[Dict[str, Any]]:
    instance, selected_k = infer_instance_and_k(file_name)
    dp = data.get("dp_internal_summary", {}) or {}
    rows: List[Dict[str, Any]] = []
    for metric in BY_LEG_METRICS:
        metric_dict = dp.get(metric, {}) or {}
        metric_total = dp.get(TOTAL_FIELDS.get(metric, ""), None)
        for leg, count in metric_dict.items():
            start_node, end_node = split_leg(str(leg))
            rows.append({
                "instance": instance,
                "selected_k": selected_k,
                "file_name": file_name,
                "metric_name": metric,
                "leg": leg,
                "start_node": start_node,
                "end_node": end_node,
                "count": count,
                "metric_total_reported": metric_total,
                "share_of_metric_total": (count / metric_total) if isinstance(metric_total, (int, float)) and metric_total else None,
            })
    return rows


def selected_run_row(file_name: str, data: Dict[str, Any]) -> Dict[str, Any]:
    instance, selected_k = infer_instance_and_k(file_name)
    b = data.get("bounds", {}) or {}
    dp = data.get("dp_internal_summary", {}) or {}
    route = data.get("route_dp_summary", {}) or {}
    opt = data.get("optimality_cut_summary", {}) or {}
    feas = data.get("feasibility_cut_summary", {}) or {}
    subtour = data.get("subtour_cut_summary", {}) or {}
    return {
        "instance": instance,
        "selected_k": selected_k,
        "file_name": file_name,
        "lower_bound": b.get("lower_bound"),
        "upper_bound": b.get("upper_bound"),
        "relative_gap": b.get("relative_gap"),
        "runtime_seconds": b.get("runtime_seconds"),
        "node_count": b.get("node_count"),
        "snapshots": b.get("snapshots"),
        "arrival_label_dominated": dp.get("arrival_label_dominated"),
        "station_label_dominated": dp.get("station_label_dominated"),
        "energy_infeasible_transitions": dp.get("energy_infeasible_transitions"),
        "destination_time_window_rejections": dp.get("destination_time_window_rejections"),
        "station_horizon_rejections": dp.get("station_horizon_rejections"),
        "leg_frontier_calls": dp.get("leg_frontier_calls"),
        "leg_frontier_no_arrivals": dp.get("leg_frontier_no_arrivals"),
        "leg_frontier_success": dp.get("leg_frontier_success"),
        "route_dp_total_calls": route.get("total_calls"),
        "route_dp_success": route.get("success"),
        "route_dp_failure": route.get("failure"),
        "optimality_cuts_total": opt.get("total"),
        "feasibility_cuts_total": feas.get("total"),
        "subtour_cuts": subtour.get("subtour_cuts"),
        "objective_source_field": "bounds.upper_bound / solver_bounds.best_obj",
        "objective_sense": data.get("objective_sense"),
        "schema_version": data.get("schema_version"),
    }


def write_formatted_excel(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.auto_filter.ref = ws.dimensions
        for i, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 70))
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(max_len + 2, 45))
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=".", help="Folder containing callback JSON files")
    parser.add_argument("--output", default="callback_json_extraction_v2_populated.xlsx")
    parser.add_argument("--top-n", type=int, default=10, help="Top legs per instance/metric")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    all_run_rows: List[Dict[str, Any]] = []
    all_leg_rows: List[Dict[str, Any]] = []
    missing: List[str] = []

    for file_name in SELECTED_FILES:
        p = input_dir / file_name
        if not p.exists():
            missing.append(file_name)
            continue
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        all_run_rows.append(selected_run_row(file_name, data))
        all_leg_rows.extend(flatten_by_leg(file_name, data))

    selected_runs = pd.DataFrame(all_run_rows)
    leg_metrics_long = pd.DataFrame(all_leg_rows)

    if leg_metrics_long.empty:
        start_node_aggregates = pd.DataFrame(columns=["instance", "selected_k", "metric_name", "start_node", "count_sum", "share_within_instance_metric"])
        end_node_aggregates = pd.DataFrame(columns=["instance", "selected_k", "metric_name", "end_node", "count_sum", "share_within_instance_metric"])
        top_legs = pd.DataFrame(columns=["instance", "selected_k", "metric_name", "rank", "leg", "start_node", "end_node", "count", "share_within_instance_metric"])
    else:
        totals = (
            leg_metrics_long.groupby(["instance", "selected_k", "metric_name"], dropna=False)["count"]
            .sum()
            .rename("metric_sum_from_legs")
            .reset_index()
        )
        leg_metrics_long = leg_metrics_long.merge(totals, on=["instance", "selected_k", "metric_name"], how="left")
        leg_metrics_long["share_within_instance_metric"] = leg_metrics_long["count"] / leg_metrics_long["metric_sum_from_legs"]

        start_node_aggregates = (
            leg_metrics_long.groupby(["instance", "selected_k", "metric_name", "start_node"], dropna=False)["count"]
            .sum()
            .rename("count_sum")
            .reset_index()
            .merge(totals, on=["instance", "selected_k", "metric_name"], how="left")
        )
        start_node_aggregates["share_within_instance_metric"] = start_node_aggregates["count_sum"] / start_node_aggregates["metric_sum_from_legs"]
        start_node_aggregates = start_node_aggregates.sort_values(["instance", "metric_name", "count_sum"], ascending=[True, True, False])

        end_node_aggregates = (
            leg_metrics_long.groupby(["instance", "selected_k", "metric_name", "end_node"], dropna=False)["count"]
            .sum()
            .rename("count_sum")
            .reset_index()
            .merge(totals, on=["instance", "selected_k", "metric_name"], how="left")
        )
        end_node_aggregates["share_within_instance_metric"] = end_node_aggregates["count_sum"] / end_node_aggregates["metric_sum_from_legs"]
        end_node_aggregates = end_node_aggregates.sort_values(["instance", "metric_name", "count_sum"], ascending=[True, True, False])

        top_legs = (
            leg_metrics_long.sort_values(["instance", "metric_name", "count"], ascending=[True, True, False])
            .groupby(["instance", "selected_k", "metric_name"], as_index=False, group_keys=False)
            .head(args.top_n)
            .copy()
        )
        top_legs["rank"] = top_legs.groupby(["instance", "selected_k", "metric_name"])["count"].rank(method="first", ascending=False).astype(int)
        top_legs = top_legs[["instance", "selected_k", "metric_name", "rank", "leg", "start_node", "end_node", "count", "share_within_instance_metric", "file_name"]]

    validation = []
    if not leg_metrics_long.empty:
        for _, row in leg_metrics_long.groupby(["instance", "metric_name"], as_index=False)["count"].sum().iterrows():
            validation.append({
                "check": "leg metric sum computed",
                "instance": row["instance"],
                "metric_name": row["metric_name"],
                "value": row["count"],
            })
    for m in missing:
        validation.append({"check": "missing selected JSON file", "instance": "", "metric_name": "", "value": m})

    notes = pd.DataFrame([
        {"item": "double_counting_note", "detail": "start_node_aggregates and end_node_aggregates are sums of leg-event counts. They are not unique labels or unique DP states."},
        {"item": "excluded_instances", "detail": "C101C6_3 and C101C6_4 are excluded by instruction."},
        {"item": "selected_run_convention", "detail": "For near-ties at numerical precision, lower K_max is preferred."},
    ])

    sheets = {
        "selected_runs": selected_runs,
        "leg_metrics_long": leg_metrics_long,
        "start_node_aggregates": start_node_aggregates,
        "end_node_aggregates": end_node_aggregates,
        "top_legs": top_legs,
        "validation": pd.DataFrame(validation),
        "notes": notes,
    }
    write_formatted_excel(Path(args.output), sheets)
    print(f"Wrote {args.output}")
    if missing:
        print("Missing selected files:", ", ".join(missing))


if __name__ == "__main__":
    main()
